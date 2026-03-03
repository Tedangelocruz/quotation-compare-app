"""
Cotización Rapida — Enterprise Quotation Comparison Tool
Refactored with modular extractor architecture.
"""

import streamlit as st
import pandas as pd
import os
import time
from io import BytesIO

from utils import get_base64_image, normalize_sku
from db import (
    init_db, save_to_db, get_items_by_quotation_id,
    update_items_batch, delete_items_by_ids, get_quotation_ids_for_filename,
    get_quotation_metadata
)
from pipeline import process_file, preview_detection, get_extractor_names
from verification import verify_items

# Page Config
st.set_page_config(page_title="Cotización Rapida", page_icon="📊", layout="wide")


def inject_custom_css(theme='dark'):
    if theme == 'light':
        bg_color, secondary_bg, text_color = "#ffffff", "#f8f9fa", "#262730"
        border_color, input_bg, input_border = "rgba(0,0,0,0.08)", "#ffffff", "#d3d3d3"
        accent_primary, accent_secondary = "#2563eb", "#059669"
        shadow_color, hover_bg = "rgba(0,0,0,0.05)", "rgba(37,99,235,0.05)"
        button_text, btn_start, btn_end = "#ffffff", "#2563eb", "#059669"
    else:
        bg_color, secondary_bg, text_color = "#0e1117", "#1e2530", "#fafafa"
        border_color, input_bg, input_border = "rgba(255,255,255,0.08)", "#262730", "#4a4a4a"
        accent_primary, accent_secondary = "#60a5fa", "#34d399"
        shadow_color, hover_bg = "rgba(0,0,0,0.3)", "rgba(96,165,250,0.1)"
        button_text, btn_start, btn_end = "#ffffff", "#60a5fa", "#34d399"
    
    st.markdown(f"""<style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@400;500;600;700&display=swap');
    * {{ transition: all 0.2s ease-in-out; }}
    .stApp {{ background-color: {bg_color} !important; color: {text_color} !important; font-family: 'Poppins', sans-serif !important; }}
    section[data-testid="stSidebar"] {{ background: linear-gradient(180deg, {secondary_bg} 0%, {bg_color} 100%) !important; border-right: 1px solid {border_color}; }}
    section[data-testid="stSidebar"] * {{ color: {text_color} !important; }}
    .block-container {{ padding-top: 2rem; padding-bottom: 3rem; max-width: 95% !important; }}
    p, span, div, label, li, td, th {{ color: {text_color} !important; }}
    h1, h2, h3, h4, h5, h6 {{ font-family: 'Poppins', sans-serif !important; font-weight: 600 !important; color: {text_color} !important; }}
    h1 {{ background: linear-gradient(135deg, {accent_primary}, {accent_secondary}); -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text; }}
    input, textarea, select {{ background-color: {input_bg} !important; color: {text_color} !important; border: 2px solid {input_border} !important; border-radius: 8px !important; padding: 10px !important; font-family: 'Poppins', sans-serif !important; }}
    input:focus, textarea:focus, select:focus {{ border-color: {accent_primary} !important; box-shadow: 0 0 0 3px {hover_bg} !important; }}
    .stButton button {{ background: linear-gradient(135deg, {btn_start}, {btn_end}) !important; color: {button_text} !important; border: none !important; border-radius: 10px !important; padding: 12px 28px !important; font-weight: 600 !important; box-shadow: 0 4px 12px {shadow_color} !important; }}
    .stButton button:hover {{ transform: translateY(-2px) !important; box-shadow: 0 6px 20px {shadow_color} !important; }}
    div[data-testid="stMetric"] {{ background: linear-gradient(135deg, {secondary_bg}, {bg_color}) !important; padding: 24px !important; border-radius: 16px !important; border: 1px solid {border_color} !important; box-shadow: 0 8px 16px {shadow_color} !important; position: relative !important; overflow: hidden !important; }}
    div[data-testid="stMetric"]::before {{ content: ''; position: absolute; top: 0; left: 0; width: 4px; height: 100%; background: linear-gradient(180deg, {accent_primary}, {accent_secondary}); }}
    div[data-testid="stMetric"]:hover {{ transform: translateY(-4px) !important; box-shadow: 0 12px 24px {shadow_color} !important; }}
    div[data-testid="stMetricLabel"] {{ font-size: 0.95rem !important; opacity: 0.8 !important; font-weight: 500 !important; text-transform: uppercase !important; letter-spacing: 0.5px !important; }}
    div[data-testid="stMetricValue"] {{ font-size: 2.2rem !important; font-weight: 700 !important; margin-top: 8px !important; }}
    .stTabs [data-baseweb="tab-list"] {{ gap: 8px; background-color: {secondary_bg}; padding: 8px; border-radius: 12px; }}
    .stTabs [data-baseweb="tab"] {{ height: 50px; background-color: transparent; border-radius: 8px; padding: 12px 24px; color: {text_color} !important; font-weight: 500 !important; border: none !important; }}
    .stTabs [aria-selected="true"] {{ background: linear-gradient(135deg, {btn_start}, {btn_end}) !important; color: white !important; }}
    [data-testid="stDataFrame"] {{ border-radius: 12px !important; overflow: hidden !important; box-shadow: 0 4px 12px {shadow_color} !important; }}
    [data-testid="stDataFrame"] th {{ background: linear-gradient(135deg, {secondary_bg}, {bg_color}) !important; padding: 16px !important; font-weight: 600 !important; border-bottom: 2px solid {accent_primary} !important; }}
    [data-testid="stFileUploader"] {{ border: 2px dashed {border_color} !important; border-radius: 16px !important; padding: 32px !important; background: linear-gradient(135deg, {secondary_bg}, {bg_color}) !important; }}
    [data-testid="stFileUploader"]:hover {{ border-color: {accent_primary} !important; }}
    [data-testid="stFileUploader"] label, [data-testid="stFileUploader"] span, [data-testid="stFileUploader"] p {{ color: {text_color} !important; }}
    [data-testid="stFileUploader"] section {{ background-color: transparent !important; color: {text_color} !important; }}
    .stProgress > div > div > div {{ background: linear-gradient(90deg, {accent_primary}, {accent_secondary}) !important; border-radius: 10px !important; }}
    hr {{ border: none !important; height: 2px !important; background: linear-gradient(90deg, transparent, {border_color}, transparent) !important; margin: 24px 0 !important; }}
    ::-webkit-scrollbar {{ width: 10px; height: 10px; }}
    ::-webkit-scrollbar-track {{ background: {secondary_bg}; border-radius: 10px; }}
    ::-webkit-scrollbar-thumb {{ background: linear-gradient(180deg, {accent_primary}, {accent_secondary}); border-radius: 10px; }}
    .supplier-badge {{ display: inline-flex; align-items: center; gap: 8px; padding: 8px 16px; border-radius: 12px; font-weight: 600; font-size: 0.9rem; margin: 4px; }}
    .badge-deterministic {{ background: linear-gradient(135deg, rgba(16,185,129,0.15), rgba(5,150,105,0.15)); border: 1px solid rgba(16,185,129,0.4); }}
    .badge-ai {{ background: linear-gradient(135deg, rgba(251,191,36,0.15), rgba(245,158,11,0.15)); border: 1px solid rgba(251,191,36,0.4); }}
    .badge-manual {{ background: linear-gradient(135deg, rgba(96,165,250,0.15), rgba(59,130,246,0.15)); border: 1px solid rgba(96,165,250,0.4); }}
    .verify-pass {{ color: #10b981; font-weight: 600; }}
    .verify-warn {{ color: #f59e0b; font-weight: 600; }}
    .verify-fail {{ color: #ef4444; font-weight: 600; }}
    </style>""", unsafe_allow_html=True)

inject_custom_css()
init_db()

# --- Session State ---
if 'session_items' not in st.session_state:
    st.session_state.session_items = []
if 'processed_files' not in st.session_state:
    st.session_state.processed_files = set()
if 'detection_results' not in st.session_state:
    st.session_state.detection_results = {}
if 'theme' not in st.session_state:
    st.session_state.theme = 'dark'

inject_custom_css(st.session_state.theme)

def clear_session():
    st.session_state.session_items = []
    st.session_state.processed_files = set()
    st.session_state.detection_results = {}
    st.rerun()

# --- Header ---
st.markdown(f"""
    <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@400;600;700&display=swap" rel="stylesheet">
    <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 2rem;">
        <div style="flex: 0 0 200px;">
            <img src="data:image/png;base64,{get_base64_image('assets/logo.png') if os.path.exists('assets/logo.png') else ''}" style="width: 200px; border-radius: 10px;" />
        </div>
        <div style="flex: 1; text-align: center;">
            <img src="data:image/png;base64,{get_base64_image('assets/title_logo.png') if os.path.exists('assets/title_logo.png') else ''}" style="max-width: 400px; height: auto;" />
        </div>
        <div style="flex: 0 0 200px; text-align: right;">
            <img src="data:image/png;base64,{get_base64_image('assets/header_icon.png') if os.path.exists('assets/header_icon.png') else ''}" style="width: 120px;" />
        </div>
    </div>
""", unsafe_allow_html=True)

# Theme toggle
col_theme1, col_theme2 = st.columns([10, 1])
with col_theme2:
    theme_icon = "🌙" if st.session_state.theme == 'dark' else "☀️"
    if st.button(theme_icon, help=f"Switch to {'Light' if st.session_state.theme == 'dark' else 'Dark'} Mode", key="theme_toggle"):
        st.session_state.theme = 'light' if st.session_state.theme == 'dark' else 'dark'
        st.rerun()

# --- Sidebar ---
with st.sidebar:
    st.header("Session")
    if st.button("🆕 Start New Quotation", type="primary"):
        clear_session()
    
    st.divider()
    st.header("Settings")
    
    exchange_rate = st.number_input("💱 Exchange Rate (USD to DOP)", value=60.0, min_value=1.0, step=0.1, format="%.2f")
    st.info(f"1 USD = {exchange_rate} DOP")
    
    api_key = st.text_input("Gemini API Key (Optional)", type="password", help="For AI-powered extraction of unknown suppliers")
    
    st.divider()
    st.markdown("### 🏗️ Architecture")
    st.markdown("""
    - 🟢 **Deterministic** — Known supplier
    - 🟡 **AI (Gemini)** — Unknown supplier  
    - 🔵 **Manual** — User override
    """)

# --- File Upload (PDF + Excel) ---
uploaded_files = st.file_uploader(
    "Upload Quotations (PDF or Excel)", 
    type=["pdf", "xlsx", "xls"], 
    accept_multiple_files=True
)

if uploaded_files:
    new_files = [f for f in uploaded_files if f.name not in st.session_state.processed_files]
    
    if new_files:
        # --- Supplier Detection Badges ---
        st.markdown("### 🔍 Supplier Detection")
        
        for f in new_files:
            if f.name not in st.session_state.detection_results:
                f.seek(0)
                detections = preview_detection(f, f.name)
                f.seek(0)
                st.session_state.detection_results[f.name] = detections
            
            detections = st.session_state.detection_results[f.name]
            best = detections[0] if detections else None
            
            if best:
                badge_class = f"badge-{best['type']}"
                st.markdown(
                    f'<div class="supplier-badge {badge_class}">'
                    f'📄 <strong>{f.name}</strong> → {best["badge"]}'
                    f'</div>',
                    unsafe_allow_html=True
                )
            else:
                st.markdown(f'<div class="supplier-badge badge-manual">📄 <strong>{f.name}</strong> → 🔴 No extractor available</div>', unsafe_allow_html=True)
        
        st.markdown("---")
        
        # --- Verification Flow ---
        if 'verification_index' not in st.session_state:
            st.session_state.verification_index = 0
        if 'file_hints' not in st.session_state:
            st.session_state.file_hints = {}
        if 'file_currencies' not in st.session_state:
            st.session_state.file_currencies = {}
        if 'file_taxes' not in st.session_state:
            st.session_state.file_taxes = {}
        if 'file_discounts' not in st.session_state:
            st.session_state.file_discounts = {}
        if 'file_overrides' not in st.session_state:
            st.session_state.file_overrides = {}
        
        current_idx = st.session_state.verification_index
        
        if current_idx < len(new_files):
            preview_file = new_files[current_idx]
            detections = st.session_state.detection_results.get(preview_file.name, [])
            best_det = detections[0] if detections else None
            is_deterministic = best_det and best_det['type'] == 'deterministic'
            
            with st.expander(f"🔍 Verify File {current_idx + 1}/{len(new_files)}: {preview_file.name}", expanded=True):
                # Show detection result
                if best_det:
                    st.markdown(f"**Detected:** {best_det['badge']}")
                
                # Extractor Override
                available = get_extractor_names()
                default_idx = 0
                if best_det and best_det['name'] in available:
                    default_idx = available.index(best_det['name'])
                
                selected_extractor = st.selectbox(
                    "🔧 Extractor Override",
                    available,
                    index=default_idx,
                    key=f"override_{current_idx}",
                    help="Override auto-detection if needed"
                )
                st.session_state.file_overrides[preview_file.name] = selected_extractor
                
                # Currency, Tax, Discount
                col_c, col_t, col_d = st.columns(3)
                with col_c:
                    curr = st.session_state.file_currencies.get(preview_file.name, "DOP")
                    file_currency = st.radio("💵 Currency:", ["DOP", "USD"], index=0 if curr == "DOP" else 1, key=f"currency_{current_idx}", horizontal=True)
                    st.session_state.file_currencies[preview_file.name] = file_currency
                with col_t:
                    file_tax = st.number_input("Tax %", min_value=0.0, max_value=100.0, value=float(st.session_state.file_taxes.get(preview_file.name, 0.0)), step=0.1, key=f"tax_{current_idx}")
                    st.session_state.file_taxes[preview_file.name] = file_tax
                with col_d:
                    file_discount = st.number_input("Discount %", min_value=0.0, max_value=100.0, value=float(st.session_state.file_discounts.get(preview_file.name, 0.0)), step=0.1, key=f"discount_{current_idx}")
                    st.session_state.file_discounts[preview_file.name] = file_discount
                
                # Hints (only for AI/fallback extractors)
                if not is_deterministic:
                    st.markdown("**Example Values (Optional):** Help the AI by providing the FIRST value from each column.")
                    current_hints = st.session_state.file_hints.get(preview_file.name, {})
                    ch1, ch2, ch3, ch4 = st.columns(4)
                    with ch1:
                        sku_hint = st.text_input("First SKU", value=current_hints.get('sku', ''), key=f"hint_sku_{current_idx}")
                    with ch2:
                        item_hint = st.text_input("First Item", value=current_hints.get('item', ''), key=f"hint_item_{current_idx}")
                    with ch3:
                        qty_hint = st.text_input("First Qty", value=current_hints.get('qty', ''), key=f"hint_qty_{current_idx}")
                    with ch4:
                        price_hint = st.text_input("First Price", value=current_hints.get('price', ''), key=f"hint_price_{current_idx}")
                    st.session_state.file_hints[preview_file.name] = {'sku': sku_hint, 'item': item_hint, 'qty': qty_hint, 'price': price_hint}
                
                # Test & Confirm buttons
                col_btn1, col_btn2 = st.columns(2)
                with col_btn1:
                    if st.button(f"🧪 Test Extract", key=f"test_{current_idx}"):
                        try:
                            preview_file.seek(0)
                            result = process_file(
                                preview_file, preview_file.name,
                                api_key=api_key,
                                example_hints=st.session_state.file_hints.get(preview_file.name, {}),
                                override_extractor=st.session_state.file_overrides.get(preview_file.name),
                            )
                            if result.items:
                                st.success(f"✅ Found {len(result.items)} items")
                                st.markdown(f"**Extractor:** {result.badge_text}")
                                verify = verify_items(result.items)
                                css_class = f"verify-{'pass' if verify.status_emoji == '✅' else 'warn' if verify.status_emoji == '⚠️' else 'fail'}"
                                st.markdown(f'<span class="{css_class}">{verify.status_emoji} {verify.status_text}</span>', unsafe_allow_html=True)
                                preview_df = pd.DataFrame(result.items)
                                show_cols = [c for c in ['product_id', 'product_name', 'quantity', 'unit_price', 'total_price', 'equipment'] if c in preview_df.columns]
                                st.dataframe(preview_df[show_cols])
                            else:
                                st.warning("No items extracted. Try adding hints or changing the extractor.")
                        except Exception as e:
                            st.error(f"Test failed: {e}")
                
                with col_btn2:
                    if st.button("✅ Confirm & Next", key=f"confirm_{current_idx}", type="primary"):
                        st.session_state.verification_index += 1
                        st.rerun()
            
            if st.button("⏩ Skip Verification & Process All", key="skip_verification"):
                st.session_state.verification_index = len(new_files)
                st.rerun()
        
        else:
            # All verified — Process
            st.success(f"✅ All {len(new_files)} files verified! Ready to process.")
            
            if st.button("🚀 Process All Files", type="primary"):
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                for idx, uploaded_file in enumerate(new_files):
                    status_text.text(f"Processing {uploaded_file.name}...")
                    try:
                        uploaded_file.seek(0)
                        result = process_file(
                            uploaded_file, uploaded_file.name,
                            api_key=api_key,
                            example_hints=st.session_state.file_hints.get(uploaded_file.name, {}),
                            override_extractor=st.session_state.file_overrides.get(uploaded_file.name),
                        )
                        
                        if result.items:
                            default_supplier = os.path.splitext(uploaded_file.name)[0]
                            for item in result.items:
                                if not item.get('supplier_name') or item['supplier_name'] == "Unknown Supplier":
                                    item['supplier_name'] = default_supplier
                            
                            file_currency = st.session_state.file_currencies.get(uploaded_file.name, result.currency or "DOP")
                            file_tax = st.session_state.file_taxes.get(uploaded_file.name, 0.0)
                            file_discount = st.session_state.file_discounts.get(uploaded_file.name, 0.0)
                            
                            # Verification status
                            v = result.verification
                            v_status = v.status.value if hasattr(v.status, 'value') else str(v.status)
                            
                            q_id = save_to_db(
                                uploaded_file.name, result.items,
                                currency=file_currency, tax_rate=file_tax, discount_rate=file_discount,
                                supplier_detected=result.supplier_name,
                                extractor_used=result.extractor_name,
                                detection_confidence=result.confidence,
                                verification_status=v_status,
                            )
                            
                            new_db_items = get_items_by_quotation_id(q_id)
                            st.session_state.session_items.extend(new_db_items)
                            st.session_state.processed_files.add(uploaded_file.name)
                            st.toast(f"✅ {len(result.items)} items from {uploaded_file.name} ({result.badge_text})")
                        else:
                            st.warning(f"Could not extract items from {uploaded_file.name}")
                    except Exception as e:
                        st.error(f"Error processing {uploaded_file.name}: {e}")
                    
                    progress_bar.progress((idx + 1) / len(new_files))
                
                status_text.text("Processing complete!")
                st.session_state.verification_index = 0
                st.session_state.file_hints = {}
                st.session_state.file_taxes = {}
                st.session_state.file_discounts = {}
                st.session_state.file_overrides = {}
                st.session_state.detection_results = {}
                st.rerun()

# --- Results Area ---
if st.session_state.session_items:
    st.divider()
    df = pd.DataFrame(st.session_state.session_items)
    total_items = len(df)
    total_suppliers = df['supplier_name'].nunique() if 'supplier_name' in df.columns else 0
    total_spend = df['total_price'].sum() if 'total_price' in df.columns else 0.0
    
    potential_savings = 0.0
    if 'sku' in df.columns and not df.empty:
        valid_skus = df[df['sku'].astype(str).str.len() >= 6].copy()
        if not valid_skus.empty:
            sku_stats = valid_skus.groupby('sku')['total_price'].agg(['min', 'max'])
            potential_savings = (sku_stats['max'] - sku_stats['min']).sum()

    col_m1, col_m2, col_m3, col_m4 = st.columns(4)
    col_m1.metric("Total Items", total_items)
    col_m2.metric("Suppliers", total_suppliers)
    col_m3.metric("Total Value (Gross)", f"${total_spend:,.2f}")
    col_m4.metric("Potential Savings", f"${potential_savings:,.2f}")

    st.markdown("---")

    tab_review, tab_compare, tab_export = st.tabs(["📝 Data Review", "📊 Price Comparison", "📤 Export"])

    # --- Tab 1: Data Review ---
    with tab_review:
        st.subheader("Review Extracted Data")
        
        # Show extraction method badges
        if 'extractor_used' in df.columns:
            extractors_used = df[['supplier_name', 'extractor_used', 'detection_confidence', 'verification_status']].drop_duplicates()
            for _, row in extractors_used.iterrows():
                ext_type = 'deterministic' if row.get('extractor_used') not in ['AI (Gemini)', 'manual'] else ('ai' if row.get('extractor_used') == 'AI (Gemini)' else 'manual')
                badge_cls = f"badge-{ext_type}"
                conf = int(float(row.get('detection_confidence', 0)) * 100)
                v_status = row.get('verification_status', 'pending')
                v_class = f"verify-{'pass' if v_status == 'pass' else 'warn' if v_status == 'warn' else 'fail'}"
                v_emoji = '✅' if v_status == 'pass' else '⚠️' if v_status == 'warn' else '❌' if v_status == 'fail' else '⏳'
                
                st.markdown(
                    f'<div class="supplier-badge {badge_cls}">'
                    f'{row["supplier_name"]} — {row.get("extractor_used", "?")} ({conf}%)'
                    f' | <span class="{v_class}">{v_emoji} {v_status}</span>'
                    f'</div>',
                    unsafe_allow_html=True
                )
        
        editor_df = df.drop(columns=['quotation_id', 'supplier_detected', 'extractor_used', 'detection_confidence', 'verification_status'], errors='ignore')

        def calculate_net_price(row):
            price = row.get('unit_price', 0) or 0
            discount = row.get('discount_rate', 0.0) or 0
            tax = row.get('tax_rate', 0.0) or 0
            return price * (1 - discount/100.0) * (1 + tax/100.0)
        
        editor_df['net_price'] = editor_df.apply(calculate_net_price, axis=1)

        edited_df = st.data_editor(
            editor_df,
            column_config={
                "id": None,
                "supplier_name": "Supplier",
                "product_name": "Product",
                "sku": "Product ID",
                "quantity": st.column_config.NumberColumn("Qty", format="%.2f"),
                "unit_price": st.column_config.NumberColumn("Price (Gross)", format="$%.2f"),
                "total_price": st.column_config.NumberColumn("Total (Gross)", format="$%.2f"),
                "net_price": st.column_config.NumberColumn("Net Price", format="$%.2f", disabled=True),
                "tax_rate": st.column_config.NumberColumn("Tax %", format="%.1f%%", disabled=True),
                "discount_rate": st.column_config.NumberColumn("Disc %", format="%.1f%%", disabled=True),
                "currency": st.column_config.TextColumn("Curr", disabled=True),
                "equipment": st.column_config.TextColumn("Equipment", disabled=True),
            },
            use_container_width=True, hide_index=True, num_rows="dynamic", key="data_editor"
        )

        if st.button("💾 Save Changes", type="primary"):
            if update_items_batch(edited_df):
                st.toast("✅ Changes saved!")
                st.session_state.session_items = edited_df.to_dict('records')
                st.rerun()
    
    # --- Tab 2: Price Comparison ---
    with tab_compare:
        st.subheader("Price Comparison")
        
        if not df.empty:
            df['sku'] = df['sku'].fillna('')
            df['product_name'] = df['product_name'].fillna('Unknown Product')
            df_filtered = df[df['sku'].astype(str).str.len() >= 6].copy()
            
            if df_filtered.empty:
                st.warning("No items with SKUs of 6+ characters found.")
            else:
                df_filtered['normalized_sku'] = df_filtered['sku'].apply(normalize_sku)
                
                def find_base_sku(norm_sku, all_norm):
                    candidates = [norm_sku]
                    for other in all_norm:
                        if not other: continue
                        if norm_sku in other or other in norm_sku:
                            candidates.append(other)
                    return min(candidates, key=len) if candidates else norm_sku
                
                all_norm = df_filtered['normalized_sku'].unique()
                df_filtered['base_sku'] = df_filtered['normalized_sku'].apply(lambda x: find_base_sku(x, all_norm))
                df_sorted = df_filtered.sort_values(['base_sku', 'supplier_name'])
                
                def convert_price_to_dop(row):
                    price = row['unit_price']
                    d_rate = row.get('discount_rate', 0.0) or 0
                    t_rate = row.get('tax_rate', 0.0) or 0
                    price = price * (1 - d_rate/100.0) * (1 + t_rate/100.0)
                    if row.get('currency') == 'USD':
                        price *= exchange_rate
                    return price

                def convert_total_to_dop(row):
                    total = row['total_price']
                    d_rate = row.get('discount_rate', 0.0) or 0
                    t_rate = row.get('tax_rate', 0.0) or 0
                    total = total * (1 - d_rate/100.0) * (1 + t_rate/100.0)
                    if row.get('currency') == 'USD':
                        total *= exchange_rate
                    return total
                
                df_sorted['unit_price_dop'] = df_sorted.apply(convert_price_to_dop, axis=1)
                df_sorted['total_price_dop'] = df_sorted.apply(convert_total_to_dop, axis=1)
                
                display_rows = []
                grouped = df_sorted.groupby('base_sku', sort=False)
                for i, (base_sku, group) in enumerate(grouped):
                    if not base_sku: continue
                    for _, row in group.iterrows():
                        display_rows.append({
                            'id': row['id'], 'quotation_id': row['quotation_id'],
                            'SKU': row['sku'], 'Product Name': row['product_name'],
                            'Supplier Name': row['supplier_name'],
                            'Currency': row.get('currency', 'DOP'),
                            'Quantity': row['quantity'],
                            'Unit Price (Final)': row['unit_price_dop'],
                            'Total (Final)': row['total_price_dop'],
                            'Tax Rate': f"{row.get('tax_rate', 0.0)}%",
                            'Discount Rate': f"{row.get('discount_rate', 0.0)}%",
                        })
                    display_rows.append({k: None if k != 'SKU' else '' for k in ['id','quotation_id','SKU','Product Name','Supplier Name','Currency','Quantity','Unit Price (Final)','Total (Final)','Tax Rate','Discount Rate']})
                
                if display_rows and display_rows[-1].get('SKU') == '':
                    display_rows.pop()
                
                comparison_display_df = pd.DataFrame(display_rows)
                
                st.info(f"💡 Prices include tax/discount. All in **DOP** (1 USD = {exchange_rate} DOP).")
                
                edited_comparison = st.data_editor(
                    comparison_display_df, key="price_comparison_editor",
                    num_rows="dynamic", use_container_width=True, hide_index=True,
                    column_config={
                        "id": None, "quotation_id": None,
                        "SKU": st.column_config.TextColumn("SKU", width="medium"),
                        "Product Name": st.column_config.TextColumn("Product Name", width="large"),
                        "Supplier Name": st.column_config.TextColumn("Supplier Name", width="large"),
                        "Currency": st.column_config.TextColumn("Orig. Currency", width="small"),
                        "Quantity": st.column_config.NumberColumn("Quantity", width="small"),
                        "Unit Price (Final)": st.column_config.NumberColumn("Unit Price (Final)", format="%.2f", width="medium"),
                        "Total (Final)": st.column_config.NumberColumn("Total (Final)", format="%.2f", width="medium"),
                        "Tax Rate": st.column_config.TextColumn("Tax", width="small"),
                        "Discount Rate": st.column_config.TextColumn("Discount", width="small"),
                    }
                )
                
                if st.button("💾 Save Comparison Changes", type="primary", key="save_comparison"):
                    original_ids = set(comparison_display_df[comparison_display_df['id'].notna()]['id'])
                    current_ids = set(edited_comparison[edited_comparison['id'].notna()]['id'])
                    deleted_ids = original_ids - current_ids
                    
                    if deleted_ids:
                        delete_items_by_ids(deleted_ids)
                        st.toast(f"Deleted {len(deleted_ids)} items")
                    
                    updates_df = edited_comparison[edited_comparison['id'].notna()].copy()
                    updates_df = updates_df.rename(columns={'SKU': 'sku', 'Product Name': 'product_name', 'Supplier Name': 'supplier_name', 'Quantity': 'quantity'})
                    
                    def reverse_price(row):
                        p = row['Unit Price (Final)']
                        try:
                            t = float(str(row.get('Tax Rate', '0')).replace('%', ''))
                            d = float(str(row.get('Discount Rate', '0')).replace('%', ''))
                        except: t, d = 0, 0
                        if row.get('Currency') == 'USD' and exchange_rate > 0: p /= exchange_rate
                        p /= (1 + t/100)
                        if d < 100: p /= (1 - d/100)
                        return p
                    
                    def reverse_total(row):
                        t_val = row['Total (Final)']
                        try:
                            t = float(str(row.get('Tax Rate', '0')).replace('%', ''))
                            d = float(str(row.get('Discount Rate', '0')).replace('%', ''))
                        except: t, d = 0, 0
                        if row.get('Currency') == 'USD' and exchange_rate > 0: t_val /= exchange_rate
                        t_val /= (1 + t/100)
                        if d < 100: t_val /= (1 - d/100)
                        return t_val
                    
                    updates_df['unit_price'] = updates_df.apply(reverse_price, axis=1)
                    updates_df['total_price'] = updates_df.apply(reverse_total, axis=1)
                    update_items_batch(updates_df)
                    
                    all_saved = []
                    for fn in st.session_state.processed_files:
                        for q_id in get_quotation_ids_for_filename(fn):
                            all_saved.extend(get_items_by_quotation_id(q_id))
                    st.session_state.session_items = all_saved
                    st.toast("✅ Changes saved!")
                    st.rerun()

    # --- Tab 3: Export ---
    with tab_export:
        st.subheader("Export Reports")
        col1, col2 = st.columns(2)
        
        with col1:
            csv = edited_df.to_csv(index=False).encode('utf-8')
            st.download_button("📥 Download Combined CSV", csv, "combined_quotations.csv", "text/csv", key='download-csv')
        
        with col2:
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                if 'comparison_display_df' in locals() and not comparison_display_df.empty:
                    summary_rows = []
                    group_rows_list = []
                    
                    def get_final_p(row):
                        try: return float(row['Total (Final)'])
                        except: return float('inf')
                    
                    for _, row in comparison_display_df.iterrows():
                        is_blank = (pd.isna(row.get('Unit Price (Final)')) or row.get('Unit Price (Final)') == '') and (pd.isna(row.get('SKU')) or row.get('SKU') == '')
                        if is_blank:
                            if group_rows_list:
                                summary_rows.append(min(group_rows_list, key=get_final_p))
                                group_rows_list = []
                        else:
                            group_rows_list.append(row)
                    if group_rows_list:
                        summary_rows.append(min(group_rows_list, key=get_final_p))
                    
                    if summary_rows:
                        summary_df = pd.DataFrame(summary_rows).drop(columns=['id'], errors='ignore')
                        summary_df['Tax'] = ''
                        summary_df['Transportation Cost'] = ''
                        summary_df['Complete Price'] = ''
                        grand_total = summary_df['Total (Final)'].sum()
                        
                        summary_df.to_excel(writer, index=False, sheet_name='Summary', startrow=0)
                        ws = writer.sheets['Summary']
                        from openpyxl.styles import Font, PatternFill
                        bold = Font(bold=True)
                        lr = len(summary_df) + 2
                        ws.cell(row=lr, column=1, value="GRAND TOTAL")
                        fp_col = list(summary_df.columns).index('Total (Final)') + 1
                        ws.cell(row=lr, column=fp_col, value=grand_total)
                        for c in range(1, len(summary_df.columns) + 1):
                            ws.cell(row=lr, column=c).font = bold
                    
                    export_cdf = comparison_display_df.drop(columns=['id'], errors='ignore')
                    export_cdf.to_excel(writer, index=False, sheet_name='Price Comparison')
                    ws2 = writer.sheets['Price Comparison']
                    from openpyxl.styles import PatternFill
                    green = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    
                    grp = []
                    for ri, (_, row) in enumerate(export_cdf.iterrows(), start=2):
                        is_blank = pd.isna(row.get('Unit Price (Final)')) and (pd.isna(row.get('SKU')) or row.get('SKU') == '')
                        if is_blank:
                            if grp:
                                prices = [(r, export_cdf.iloc[r-2]['Total (Final)']) for r in grp if pd.notna(export_cdf.iloc[r-2]['Total (Final)'])]
                                if prices:
                                    mr = min(prices, key=lambda x: x[1])[0]
                                    for ci in range(1, len(export_cdf.columns) + 1):
                                        ws2.cell(row=mr, column=ci).fill = green
                            grp = []
                        else:
                            grp.append(ri)
                    if grp:
                        prices = [(r, export_cdf.iloc[r-2]['Total (Final)']) for r in grp if pd.notna(export_cdf.iloc[r-2]['Total (Final)'])]
                        if prices:
                            mr = min(prices, key=lambda x: x[1])[0]
                            for ci in range(1, len(export_cdf.columns) + 1):
                                ws2.cell(row=mr, column=ci).fill = green
                else:
                    pd.DataFrame(['No comparison data']).to_excel(writer, index=False, sheet_name='Price Comparison')
                
                # --- Sheet 3: Best Price (grouped by supplier) ---
                if 'summary_rows' in dir() and summary_rows:
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    bold_font = Font(bold=True)
                    header_font = Font(bold=True, size=13, color="FFFFFF")
                    supplier_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
                    total_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
                    grand_fill = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
                    grand_font = Font(bold=True, size=14, color="FFFFFF")
                    thin_border = Border(
                        bottom=Side(style='thin', color='CCCCCC')
                    )
                    
                    # Build a DataFrame from summary_rows and group by supplier
                    best_df = pd.DataFrame(summary_rows)
                    bp_columns = ['SKU', 'Product Name', 'Quantity', 'Unit Price (Final)', 'Total (Final)']
                    
                    # Create the sheet
                    # We write manually for full control over layout
                    ws_bp = writer.book.create_sheet('Best Price')
                    
                    # Title row
                    ws_bp.cell(row=1, column=1, value="BEST PRICE BREAKDOWN BY SUPPLIER")
                    ws_bp.cell(row=1, column=1).font = Font(bold=True, size=16)
                    ws_bp.merge_cells('A1:E1')
                    
                    current_row = 3
                    overall_grand_total = 0.0
                    supplier_totals = []  # Track for summary at end
                    
                    # Group by Supplier Name
                    suppliers = best_df.groupby('Supplier Name', sort=True)
                    
                    for supplier_name, supplier_items in suppliers:
                        if not supplier_name or pd.isna(supplier_name):
                            continue
                        
                        # Supplier header row (colored banner)
                        ws_bp.cell(row=current_row, column=1, value=f"📦 {supplier_name}")
                        for col_idx in range(1, 6):
                            ws_bp.cell(row=current_row, column=col_idx).fill = supplier_fill
                            ws_bp.cell(row=current_row, column=col_idx).font = header_font
                        ws_bp.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                        current_row += 1
                        
                        # Column headers
                        col_headers = ['SKU', 'Product Name', 'Qty', 'Unit Price (DOP)', 'Total (DOP)']
                        for ci, header in enumerate(col_headers, 1):
                            cell = ws_bp.cell(row=current_row, column=ci, value=header)
                            cell.font = bold_font
                            cell.border = thin_border
                        current_row += 1
                        
                        # Data rows
                        supplier_total = 0.0
                        item_count = 0
                        for _, item_row in supplier_items.iterrows():
                            ws_bp.cell(row=current_row, column=1, value=item_row.get('SKU', ''))
                            ws_bp.cell(row=current_row, column=2, value=item_row.get('Product Name', ''))
                            ws_bp.cell(row=current_row, column=3, value=item_row.get('Quantity', 0))
                            
                            unit_p = item_row.get('Unit Price (Final)', 0) or 0
                            total_p = item_row.get('Total (Final)', 0) or 0
                            
                            cell_up = ws_bp.cell(row=current_row, column=4, value=round(unit_p, 2))
                            cell_up.number_format = '#,##0.00'
                            cell_tp = ws_bp.cell(row=current_row, column=5, value=round(total_p, 2))
                            cell_tp.number_format = '#,##0.00'
                            
                            for ci in range(1, 6):
                                ws_bp.cell(row=current_row, column=ci).border = thin_border
                            
                            supplier_total += total_p
                            item_count += 1
                            current_row += 1
                        
                        # Supplier subtotal row
                        ws_bp.cell(row=current_row, column=3, value=f"{item_count} items")
                        ws_bp.cell(row=current_row, column=4, value="SUBTOTAL:")
                        ws_bp.cell(row=current_row, column=4).font = bold_font
                        cell_st = ws_bp.cell(row=current_row, column=5, value=round(supplier_total, 2))
                        cell_st.font = bold_font
                        cell_st.number_format = '#,##0.00'
                        for ci in range(1, 6):
                            ws_bp.cell(row=current_row, column=ci).fill = total_fill
                        
                        supplier_totals.append({'name': supplier_name, 'total': supplier_total, 'items': item_count})
                        overall_grand_total += supplier_total
                        current_row += 2  # Blank row between suppliers
                    
                    # --- Grand Total Section ---
                    current_row += 1
                    ws_bp.cell(row=current_row, column=1, value="PURCHASE SUMMARY")
                    ws_bp.cell(row=current_row, column=1).font = Font(bold=True, size=14)
                    ws_bp.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                    current_row += 1
                    
                    # Summary table headers
                    sum_headers = ['Supplier', '', 'Items', '', 'Total (DOP)']
                    for ci, h in enumerate(sum_headers, 1):
                        cell = ws_bp.cell(row=current_row, column=ci, value=h)
                        cell.font = bold_font
                        cell.border = thin_border
                    current_row += 1
                    
                    # Summary rows per supplier
                    for st_info in supplier_totals:
                        ws_bp.cell(row=current_row, column=1, value=st_info['name'])
                        ws_bp.cell(row=current_row, column=3, value=st_info['items'])
                        cell_t = ws_bp.cell(row=current_row, column=5, value=round(st_info['total'], 2))
                        cell_t.number_format = '#,##0.00'
                        for ci in range(1, 6):
                            ws_bp.cell(row=current_row, column=ci).border = thin_border
                        current_row += 1
                    
                    # Grand total row
                    total_items_count = sum(s['items'] for s in supplier_totals)
                    ws_bp.cell(row=current_row, column=1, value="GRAND TOTAL")
                    ws_bp.cell(row=current_row, column=3, value=total_items_count)
                    cell_gt = ws_bp.cell(row=current_row, column=5, value=round(overall_grand_total, 2))
                    cell_gt.number_format = '#,##0.00'
                    for ci in range(1, 6):
                        ws_bp.cell(row=current_row, column=ci).fill = grand_fill
                        ws_bp.cell(row=current_row, column=ci).font = grand_font
                    
                    # Auto-size columns
                    ws_bp.column_dimensions['A'].width = 18
                    ws_bp.column_dimensions['B'].width = 45
                    ws_bp.column_dimensions['C'].width = 12
                    ws_bp.column_dimensions['D'].width = 18
                    ws_bp.column_dimensions['E'].width = 18

                if 'quotation_id' in df.columns:
                    import sqlite3
                    conn = sqlite3.connect('quotations.db')
                    for q_id in df['quotation_id'].unique():
                        if pd.isna(q_id): continue
                        q_data = pd.read_sql_query("SELECT filename, currency FROM quotations WHERE id = ?", conn, params=(int(q_id),))
                        if q_data.empty: continue
                        fn = q_data.iloc[0]['filename']
                        currency = q_data.iloc[0]['currency']
                        base = os.path.splitext(fn)[0]
                        sn = base[:30]
                        for ch in ['[',']','*','?','/','\\',' :']:
                            sn = sn.replace(ch, '')
                        q_items = df[df['quotation_id'] == q_id].copy()
                        tax_r = q_items['tax_rate'].iloc[0] if len(q_items) > 0 else 0.0
                        disc_r = q_items['discount_rate'].iloc[0] if len(q_items) > 0 else 0.0
                        sdf = q_items[['sku','product_name','quantity','unit_price','total_price']].copy()
                        sdf.columns = ['SKU','Product Name','Quantity','Unit Price','Total']
                        sdf.to_excel(writer, index=False, sheet_name=sn)
                        wsx = writer.sheets[sn]
                        from openpyxl.styles import Font
                        bold = Font(bold=True)
                        sub = sdf['Total'].sum()
                        da = sub * (disc_r / 100.0)
                        sad = sub - da
                        ta = sad * (tax_r / 100.0)
                        ft = sad + ta
                        tr = len(sdf) + 3
                        wsx.cell(row=tr, column=4, value="Subtotal:"); wsx.cell(row=tr, column=5, value=sub)
                        wsx.cell(row=tr+1, column=4, value=f"Discount ({disc_r}%):"); wsx.cell(row=tr+1, column=5, value=-da)
                        wsx.cell(row=tr+2, column=4, value="Subtotal (excl. Tax):"); wsx.cell(row=tr+2, column=5, value=sad)
                        wsx.cell(row=tr+3, column=4, value=f"Tax ({tax_r}%):"); wsx.cell(row=tr+3, column=5, value=ta)
                        wsx.cell(row=tr+4, column=4, value=f"FINAL TOTAL ({currency}):"); wsx.cell(row=tr+4, column=5, value=ft)
                        wsx.cell(row=tr+4, column=4).font = Font(bold=True, size=12)
                        wsx.cell(row=tr+4, column=5).font = Font(bold=True, size=12)
                        if currency == 'USD':
                            fdop = ft * exchange_rate
                            wsx.cell(row=tr+5, column=4, value="Exchange Rate:"); wsx.cell(row=tr+5, column=5, value=exchange_rate)
                            wsx.cell(row=tr+6, column=4, value="FINAL TOTAL (DOP):"); wsx.cell(row=tr+6, column=5, value=fdop)
                            wsx.cell(row=tr+6, column=4).font = Font(bold=True, size=12, color="008000")
                            wsx.cell(row=tr+6, column=5).font = Font(bold=True, size=12, color="008000")
                    conn.close()
            
            ts = time.strftime("%Y%m%d-%H%M%S")
            st.download_button(f"📥 Download Excel ({ts})", buffer.getvalue(), f"quotation_comparison_{ts}.xlsx",
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key='download-excel')

else:
    st.info("Upload PDFs or Excel files to start comparing.")
